import os, sys, openpyxl
from datetime import timedelta
from dateutil.parser import parse
from openpyxl import styles
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from result import Result

"""
The data file is first read in as a list of strings, then various components
are parsed and the relevant data is stored as fields in a Result object.
Finally, the derived metrics are exported into an Excel spreadsheet.

Usage:
    Single file:
    process.py <filename>

    Folder:
    process.py -f <folder>

"""


def parse_heading(ls):
    """
    Parse the heading of a block.

    :param ls: The list of strings to parse.
    :return: The date of the experiment, the subject number,
             the duration and the variable interval.
    """
    date_line = ls.pop(0)
    ls.pop(0)  # End Date
    subj_line = ls.pop(0)
    ls.pop(0)  # Experiment
    ls.pop(0)  # Group
    ls.pop(0)  # Box
    start_line = ls.pop(0)
    end_line = ls.pop(0)
    ls = ls[20:]  # Junk
    vi_line = ls.pop(0)
    ls.pop(0)  # A:

    date = parse(date_line.split()[2], dayfirst=False)
    subj = int(subj_line.split()[1])
    start = parse(start_line.split()[2])
    end = parse(end_line.split()[2])
    dur = (end - start)
    vi = float(vi_line.split()[1])

    return date, subj, dur, vi, ls


def parse_times(ls):
    """
    Parse a block of timestamps.

    :param ls: The list of strings to parse.
    :return: The timestamps, their number, and the rest of the file.
    """
    times = []
    curr_line = ls.pop(0)
    num = round(float(curr_line.split()[1]))

    while curr_line[1] != ':':
        vals = curr_line.split()[1:]
        parsed = list(map(float, vals))
        times = times + parsed

        curr_line = ls.pop(0)

    times = times[1:]  # Remove number

    return num, times, ls


def parse_block(ls):
    """
    Parse a whole block of data, i.e. the measurements for one experiment.

    :param ls: The list of strings to parse.
    :return: The heading data and timestamps for active presses, inactive presses,
             rewards and magazine entries.
    """
    # Parse the heading of the block
    date, subj, dur, vi, ls = parse_heading(ls)

    # Active presses
    act_num, act_times, ls = parse_times(ls)

    # Inactive presses
    inact_num, inact_times, ls = parse_times(ls)

    # Rewards
    rew_num, rew_times, ls = parse_times(ls)

    # Magazine entries
    mag_num, _, ls = parse_times(ls)

    ls = ls[3:]  # T: & Z:

    if rew_num == 30:
        dur = round(timedelta(seconds=rew_times[-1]).seconds / 60.0, 2)
    else:
        dur = round(timedelta(seconds=720.0).seconds / 60.0, 2)

    return date, subj, dur, vi, act_num, act_times, inact_num, inact_times, \
        rew_num, rew_times, mag_num, ls


def find_cont_substring(l1, l2, n):
    """Dynamic programming algorithm to find continuous, uninterrupted substrings in two increasing lists.

    Given two lists, finds the position in the first one which is proceeded
    by n elements from the first list, not interrupted by the second.

    For example,
        find_cont_substring([3,5,6,7,8,10], [1,2,4,9], 4)
    would return 1, as the substring in the first list starting at index 1 ([5,6,7,8])
    has length 4 and is not interrupted by any element in the second list.
    """

    # If there are not enough data points, return the base result
    if len(l1) < n and len(l2) < n:
        return 800.0

    # Tag first list elements with True, second list elements with False
    tagged1 = [(v, True) for v in l1]
    tagged2 = [(v, False) for v in l2]

    merged = tagged1 + tagged2
    merged.sort(key=lambda pair: pair[0])

    # max_cont_ss[i] = n if n values of merged before (and including) i have the True tag
    max_cont_ss = [1 if merged[0][1] else 0]

    string_of_n = []

    for i in range(1, len(merged)):
        if merged[i][1]:  # "Streak" is continued
            max_cont_ss.append(max_cont_ss[i - 1] + 1)
        else:  # "Streak" ends
            max_cont_ss.append(0)

        if max_cont_ss[i] == n:  # If reached desired limit, save the last value
            string_of_n.append(merged[i])

    return 800.0 if len(string_of_n) == 0 else string_of_n[0][0]


def process_block(ls):
    """
    Process a block of data read from the file.

    :param ls: The list of strings to parse and process.
    :return: A Result object containing the derived measurements.
    """

    # Parse the first block of the input
    date, subj, dur, vi, act_num, act_times, inact_num, inact_times, \
        rew_num, rew_times, mag_num, ls = parse_block(ls)

    vi /= 100

    # Number of measurements within the variable interval
    num_vi_al = len([t for t in act_times if t < vi])
    num_vi_il = len([t for t in inact_times if t < vi])

    within_vi = num_vi_al + num_vi_il

    # Average response rates in the variable interval
    arr_al = round(num_vi_al / vi, 2)
    arr_il = round(num_vi_il / vi, 2)

    # Time of the first pedal press
    if act_num != 0 and inact_num != 0:
        lat_r1 = min(act_times[0], inact_times[0])
    elif act_num != 0:
        lat_r1 = act_times[0]
    elif inact_num != 0:
        lat_r1 = inact_times[0]
    else:
        lat_r1 = 800.0

    # Presses before first reward
    if len(rew_times) > 0:
        first_rew = rew_times[0]

        al_vi_to_rew = len([x for x in act_times if (x > vi / 100) and (x <= first_rew)])
        il_vi_to_rew = len([x for x in inact_times if (x > vi / 100) and (x <= first_rew)])
    else:
        al_vi_to_rew = 0
        il_vi_to_rew = 0

    # Time when the active pedal had 10 uninterrupted presses
    lat_fr10a = find_cont_substring(act_times, inact_times, 10)

    # Time when the inactive pedal had 10 uninterrupted presses
    lat_fr10i = find_cont_substring(inact_times, act_times, 10)

    # Times after the variable interval
    act_times_aft_vi = list(filter(lambda x: x > vi, act_times))
    inact_times_aft_vi = list(filter(lambda x: x > vi, inact_times))

    # Time when the active pedal had 10 uninterrupted presses after the VI
    lat_fr10a_aft_vi = find_cont_substring(act_times_aft_vi, inact_times_aft_vi, 10)

    # Time when the inactive pedal had 10 uninterrupted presses after the VI
    lat_fr10i_aft_vi = find_cont_substring(inact_times_aft_vi, act_times_aft_vi, 10)

    return Result(date, subj, dur, within_vi, act_num, inact_num, lat_r1, lat_fr10a, lat_fr10i, lat_fr10a_aft_vi,
                  lat_fr10i_aft_vi, vi, num_vi_al, arr_al, num_vi_il, arr_il, rew_num, mag_num, al_vi_to_rew,
                  il_vi_to_rew), ls


def process_file(ls):
    """
    Process a whole file.

    :param ls: The contents of a file as a list of strings (one string for each row).
    :return: A list of Result objects, one for each subject, sorted by subject number.
    """
    measurements = []
    while True:
        res, ls = process_block(ls)
        measurements.append(res)
        if len(ls) >= 3:
            ls = ls[3:]  # Whitespace after block
        else:  # No more blocks left
            break

    measurements.sort(key=lambda m: m.rat)  # Sort by rat number
    return measurements


def write_spreadsheet(ms):
    """
    Write the experiment results in a Microsoft Excel spreadsheet.

    :param ms: The list of Result objects containing the measurements.
    The formatted Excel spreadsheet is written into the file "DDT.xlsx".
    """
    wb = openpyxl.load_workbook('DDT.xlsx')
    sheet_title = ms[0].date.strftime('%d-%m-%Y')

    wb.create_sheet(title=sheet_title)
    sheet = wb.get_sheet_by_name(wb.get_sheet_names()[-1])

    # Set axis values
    sheet['A1'] = 'Rat'
    sheet['B1'] = 'Lever (FR10)'
    sheet['C1'] = 'Lever (FR10 after VI)'
    sheet['D1'] = 'Total (AL)'
    sheet['E1'] = 'Total (IL)'
    sheet['F1'] = 'Total (VI)'
    sheet['G1'] = 'Latency (R1)'
    sheet['H1'] = 'Latency (FR10A)'
    sheet['I1'] = 'Latency (FR10I)'
    sheet['J1'] = 'VI length'
    sheet['K1'] = '#AL in VI'
    sheet['L1'] = 'ARR(AL, VI)'
    sheet['M1'] = '#IL in VI'
    sheet['N1'] = 'ARR(IL, VI)'
    sheet['O1'] = 'Rewards'
    sheet['P1'] = 'Mag Entries'
    sheet['Q1'] = 'VI-Rew1 (AL)'
    sheet['R1'] = 'VI-Rew1 (IL)'
    sheet['S1'] = 'Session Time'

    # Make top row bold.
    for i in range(1, 20):
        font = Font(bold=True)
        sheet[get_column_letter(i) + '1'].font = font

    # Set row width
    sheet.column_dimensions['B'].width = 12
    sheet.column_dimensions['C'].width = 12
    sheet.column_dimensions['G'].width = 11
    sheet.column_dimensions['H'].width = 14
    sheet.column_dimensions['I'].width = 14
    sheet.column_dimensions['Q'].width = 12
    sheet.column_dimensions['R'].width = 12

    # Add rows
    for i in range(0, len(ms)):
        row = i + 2
        sheet['A' + str(row)] = ms[i].rat
        sheet['B' + str(row)] = 'A' if ms[i].lat_fr10a < ms[i].lat_fr10i else (
            'I' if ms[i].lat_fr10a > ms[i].lat_fr10i else '/')
        sheet['C' + str(row)] = 'A' if ms[i].lat_fr10a_aft_vi < ms[i].lat_fr10i_aft_vi else (
            'I' if ms[i].lat_fr10a_aft_vi > ms[i].lat_fr10i_aft_vi else '/')
        sheet['D' + str(row)] = ms[i].tot_al
        sheet['E' + str(row)] = ms[i].tot_il
        sheet['F' + str(row)] = ms[i].vp
        sheet['G' + str(row)] = ms[i].lat_r1 if ms[i].lat_r1 != 800.0 else '/'
        sheet['H' + str(row)] = ms[i].lat_fr10a if ms[i].lat_fr10a != 800.0 else '/'
        sheet['I' + str(row)] = ms[i].lat_fr10i if ms[i].lat_fr10i != 800.0 else '/'
        sheet['J' + str(row)] = ms[i].vi
        sheet['K' + str(row)] = ms[i].num_vi_al
        sheet['L' + str(row)] = ms[i].arr_al
        sheet['M' + str(row)] = ms[i].num_vi_il
        sheet['N' + str(row)] = ms[i].arr_il
        sheet['O' + str(row)] = ms[i].rew
        sheet['P' + str(row)] = ms[i].mag
        sheet['Q' + str(row)] = ms[i].al_vi_to_rew
        sheet['R' + str(row)] = ms[i].il_vi_to_rew
        sheet['S' + str(row)] = ms[i].dur

        for j in range(1, 20):
            sheet[get_column_letter(j) + str(row)].alignment = Alignment(horizontal='right')

    # Formatting
    green_fill = styles.PatternFill(start_color='99FF99', end_color='99FF99', fill_type='solid')
    red_fill = styles.PatternFill(start_color='ffc7ce', end_color='ffc7ce', fill_type='solid')

    sheet.conditional_formatting.add(
        'B2:B40',
        FormulaRule(formula=['NOT(ISERROR(SEARCH("A",B2)))'], stopIfTrue=True,
                    fill=green_fill))

    sheet.conditional_formatting.add(
        'B2:B40',
        FormulaRule(formula=['NOT(ISERROR(SEARCH("I",B2)))'], stopIfTrue=True,
                    fill=red_fill))

    sheet.conditional_formatting.add(
        'C2:C40',
        FormulaRule(formula=['NOT(ISERROR(SEARCH("A",C2)))'], stopIfTrue=True,
                    fill=green_fill))

    sheet.conditional_formatting.add(
        'C2:C40',
        FormulaRule(formula=['NOT(ISERROR(SEARCH("I",C2)))'], stopIfTrue=True,
                    fill=red_fill))

    wb.save("DDT.xlsx")


"""
Open the file or directory given as a command-line argument, and write all the
processed data in the spreadsheet.
"""
files = []

if len(sys.argv) > 2:
    directory = sys.argv[2]
    for filename in os.listdir(directory):
        if '.' not in filename:  # File without an extension
            files.append(open(os.path.join(directory, filename)))
else:
    files.append(open(sys.argv[1]))
for f in files:
    lines = f.readlines()[4:]  # Ignore header
    measurements = process_file(lines)
    write_spreadsheet(measurements)

    f.close()

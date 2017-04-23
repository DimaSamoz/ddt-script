import openpyxl
import sys
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from process import parse_block

"""
Categorise all measurements in 12 one-minute bins and save the data in a spreadshet.

"""


class Bins:
    """
    Class for the measurements of a single rat.
    """

    def __init__(self, rat, al_bins, il_bins):
        self.rat = rat
        self.al_bins = al_bins
        self.il_bins = il_bins


def process_block2(ls):
    """
    Process a block to calculate the number of presses in 1-minute bins.
    :param ls:
    :return:
    """
    date, rat, dur, vi, act_num, act_times, inact_num, inact_times, rew_num, rew_times, \
    mag_num, al_vi_to_rew, il_vi_to_rew, ls = parse_block(ls)

    al_bins = []
    il_bins = []

    for i in range(0, 12):
        al_bins.append(len([x for x in act_times if i * 60.0 <= x < (i + 1) * 60.0]))
        il_bins.append(len([x for x in inact_times if i * 60.0 <= x < (i + 1) * 60.0]))

    return Bins(rat, al_bins, il_bins), ls


def process_file2(ls):
    """
    Process the file.

    :param ls: The file contents as a list of strings.
    :return:
    """
    measurements = []
    while True:
        bins, ls = process_block2(ls)
        measurements.append(bins)
        if len(ls) >= 3:
            ls = ls[3:]  # Whitespace after block
        else:  # No more blocks left
            break

    measurements.sort(key=lambda m: m.rat)  # Sort by rat number
    return measurements


def write_spreadsheet2(ms):
    """
    Write the new spreadsheet.
    :param ms: List of measurements.
    """
    wb = openpyxl.load_workbook('DDT.xlsx')
    sheet_title = '12_min'

    wb.create_sheet(title=sheet_title)
    sheet = wb.get_sheet_by_name(wb.get_sheet_names()[-1])

    # Set axis values
    sheet['A1'] = 'Active presses'
    sheet['A2'] = 'Rat'
    sheet['B2'] = '0-1'
    sheet['C2'] = '1-2'
    sheet['D2'] = '2-3'
    sheet['E2'] = '3-4'
    sheet['F2'] = '4-5'
    sheet['G2'] = '5-6'
    sheet['H2'] = '6-7'
    sheet['I2'] = '7-8'
    sheet['J2'] = '8-9'
    sheet['K2'] = '9-10'
    sheet['L2'] = '10-11'
    sheet['M2'] = '11-12'

    sheet['A40'] = 'Inactive presses'
    sheet['A41'] = 'Rat'
    sheet['B41'] = '0-1'
    sheet['C41'] = '1-2'
    sheet['D41'] = '2-3'
    sheet['E41'] = '3-4'
    sheet['F41'] = '4-5'
    sheet['G41'] = '5-6'
    sheet['H41'] = '6-7'
    sheet['I41'] = '7-8'
    sheet['J41'] = '8-9'
    sheet['K41'] = '9-10'
    sheet['L41'] = '10-11'
    sheet['M41'] = '11-12'
    for i in range(1, 14):
        font = Font(bold=True)
        sheet[get_column_letter(i) + '2'].font = font
        sheet[get_column_letter(i) + '41'].font = font

    for i in range(0, len(ms)):
        row = i + 3
        sheet['A' + str(row)] = ms[i].rat
        for j in range(0, 12):
            col = j + 2
            sheet[get_column_letter(col) + str(row)] = ms[i].al_bins[j]

    for i in range(0, len(ms)):
        row = i + 42
        sheet['A' + str(row)] = ms[i].rat
        for j in range(0, 12):
            col = j + 2
            sheet[get_column_letter(col) + str(row)] = ms[i].il_bins[j]

        for j in range(1, 14):
            sheet[get_column_letter(j) + str(row)].alignment = Alignment(horizontal='right')

    wb.save("DDT.xlsx")


f = open(sys.argv[1])
lines = f.readlines()[4:]  # Ignore header
ms = process_file2(lines)
write_spreadsheet2(ms)

f.close()
